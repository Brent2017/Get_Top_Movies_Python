import urllib.request
import time
import re
from bs4 import BeautifulSoup
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl import load_workbook
#change for test github 20170912
#change by pc 0913

def get_one_page(url):
    response = urllib.request.urlopen(url)
    return response   

def parse_one_page(html,offset):
    wb = load_workbook(filename = 'douban_result.xlsx') #open a excel file
    ws = wb['Sheet']
    i = 1
    soup = BeautifulSoup(html,'lxml')
    for ol_table in soup.find_all(name='ol'):
        for items in ol_table.find_all(attrs={'class': 'item'}):
            i = i + 1
            row = offset + i
            cellname = 'A' + str(row)
            ws[cellname].value = int(items.find(name='em').string)
            cellname = 'C' + str(row)
            ws[cellname].value = float(items.find(attrs={'class': 'rating_num'}).string)
            cellname = 'B' + str(row)
            #ws[cellname].value = items.find(attrs={'class': 'title'}).string 
            for hd in items.find_all(attrs={'class': 'hd'}):
                ws[cellname].value = hd.find_all(name='span')[0].string + hd.find_all(name='span')[1].string
            cellname = 'D' + str(row)
            for bd in items.find_all(attrs={'class': 'bd'}):
                year = re.search('\d\d\d\d',bd.find(class_ = '').text)
                ws[cellname].value = int(year.group(0))
    wb.save(filename = 'douban_result.xlsx')
            
def main(offset):
    url = 'https://movie.douban.com/top250?start=' + str(offset) + '&filter=' #https://movie.douban.com/top250?start=50&filter=
    print(url)
    html = get_one_page(url)
    parse_one_page(html,offset)

if __name__ == '__main__':
    for i in range(10):
        main(offset=i*25)
        time.sleep(1)
