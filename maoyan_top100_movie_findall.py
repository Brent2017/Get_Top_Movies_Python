import urllib.request
import time
import re
from bs4 import BeautifulSoup
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl import load_workbook

def get_one_page(url):
    response = urllib.request.urlopen(url)
    return response   

def parse_one_page(html,offset):
    wb = load_workbook(filename = 'maoyan_result.xlsx') #open a excel file
    ws = wb['Sheet']
    i = 1
    soup = BeautifulSoup(html,'lxml')
    for content_table in soup.find_all(class_ = 'content'):   
        for items in content_table.find_all(name='dd'):          
            i = i + 1
            row = offset + i
            cellname = 'A' + str(row)
            ws[cellname].value = int(items.find(name='i').string)
            cellname = 'C' + str(row)
            ws[cellname].value = float(items.find(attrs={'class': 'integer'}).string + items.find(class_ = 'fraction').string)
            cellname = 'B' + str(row)
            ws[cellname].value = items.find(attrs={'class': 'name'}).string
            cellname = 'D' + str(row)
            for releasetime in items.find_all(attrs={'class': 'releasetime'}):
                year = re.search('\d\d\d\d',releasetime.string)
                ws[cellname].value = int(year.group(0))
    wb.save(filename = 'maoyan_result.xlsx')
            
def main(offset):
    url = 'http://maoyan.com/board/4?offset=' + str(offset)  #http://maoyan.com/board/4?offset=10
    print(url)
    html = get_one_page(url)
    parse_one_page(html,offset)

if __name__ == '__main__':
    for i in range(10):
        main(offset=i*10)
        time.sleep(1)
